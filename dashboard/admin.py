from django.contrib import admin, messages
from django.http import HttpResponse
from django.shortcuts import render
from .models import Clan, Character
from django_object_actions import DjangoObjectActions
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from .forms import ClanForm
from datetime import datetime

# Register your models here.
admin.site.site_header = "Mir4 Clan Organizer"
admin.site.site_title = "Mir4 Clan Organizer"
admin.site.site_url = "Mir4 Clan Organizer"

@admin.action(description="Fulfilled current week donation")
def fulfilled_donation(modeladmin, request, queryset):
    for query in queryset:
        query.gold_donation = 100
        query.save()

@admin.action(description="Set member as no clan")
def set_mem_no_clan(modeladmin, request, queryset):
    for query in queryset:
        query.clan = None
        query.save()

@admin.action(description="Reset every value")
def reset_donation(modeladmin, request, queryset):
    clan = queryset[0].clan
    for query in queryset:
        query.gold_donation = 0
        query.gold_debt = 0
        query.advanced_gold = 0
        query.save()
    clan.gold_donation = 0
    clan.save()

@admin.action(description="Set gold donation as per clan gold donation")
def set_donation(modeladmin, request, queryset):
    clan = queryset[0].clan
    for query in queryset:
        if clan.gold_donation != 0:
            query.gold_donation = 100
            query.save()
            clan.total_gold += 100
    clan.save()

@admin.action(description="Reset gold & calculate debts")
def reset_and_calculate(modeladmin, request, queryset):
    clan_donation = queryset[0].clan.gold_donation
    for query in queryset:
        mem_donation_total = clan_donation
        if query.advanced_gold != 0:
            query.gold_donation += query.advanced_gold 
            query.advanced_gold = 0
        
        if query.gold_debt > 0:
            mem_donation_total += query.gold_debt
            query.gold_debt = 0

        if query.gold_donation < mem_donation_total:
            difference = mem_donation_total - query.gold_donation 
            query.gold_debt += difference
            query.gold_donation = 0
            query.save()
        elif query.gold_donation == mem_donation_total:
            query.gold_donation = 0
            query.save()
        elif query.gold_donation > mem_donation_total:
            difference = query.gold_donation - mem_donation_total
            query.advanced_gold += difference
            query.gold_donation = 0
            query.save()

    queryset[0].clan.gold_donation = 0
    queryset[0].clan.save()

@admin.register(Character)
class CharacterAdmin(admin.ModelAdmin):
    list_display = [
        "name",
        "level",
        "gold_donation",
        "gold_debt",
        "advanced_gold",
        "clan"
    ]
    ordering = ['-level']
    list_filter = ['clan']
    actions = [
        reset_and_calculate,
        set_donation,
        reset_donation,
        set_mem_no_clan,
        fulfilled_donation
    ]

@admin.register(Clan)
class ClanAdmin(DjangoObjectActions, admin.ModelAdmin):
    list_display = [
        "name",
        "total_members",
        "gold_donation",
        "total_gold_donation"
    ]
    changelist_actions = ('create_report', )

    def total_gold_donation(self, obj):
        total_gold = 0
        chars = Character.objects.all().order_by('-level')
        for char in chars:
            if char.clan == obj:
                total_gold += char.gold_donation
        return total_gold
    total_gold_donation.short_description = "Total Current Gold"

    def create_report(self, request, obj):
        if 'apply' in request.POST:
            now = datetime.now()
            clan = Clan.objects.get(name=request.POST.get('clan'))
            chars = Character.objects.filter(clan=clan)
            filepath = f"media/reports/{now.month}-{now.day}-{now.year} Gold & Level Record - {clan.gold_donation}G.xlsx"

            wb = Workbook()
            wb.save(filepath)

            workbook = load_workbook(filepath)
            sheet = workbook.active

            sheet.merge_cells('A1:E1')
            sheet['A1'].value = f'DATE RECORDED: {now.month}-{now.day}-{now.year}'
            sheet['A1'].font = Font(bold=True)
            sheet['A1'].alignment = Alignment(horizontal='center')
            sheet.merge_cells('A2:E2')
            sheet['A2'].value = f'天龍门RPG GOLD DONATIONS AND LEVEL RECORD'
            sheet['A2'].font = Font(bold=True)
            sheet['A2'].alignment = Alignment(horizontal='center')
            sheet.merge_cells('A3:E3')
            sheet['A3'].value = f'GOLD DONATION: {clan.gold_donation}'
            sheet['A3'].font = Font(bold=True)
            sheet['A3'].alignment = Alignment(horizontal='center')

            AHeaders = [
                "Name",
                "Level",
                "Gold Donation",
                "Previous Gold Debt",
                "Advanced Gold Donation"
            ]

            row = 4
            for i, value in enumerate(AHeaders):
                cll = sheet.cell(
                    column = i + 1,
                    row = row,
                    value = value
                )
                cll.font = Font(bold=True)
                cll.alignment = Alignment(horizontal='center')
            row += 1

            for char in chars:
                for i, value in enumerate(AHeaders):
                    if i == 0:
                        cll = sheet.cell(column = i + 1, row = row, value = char.name)
                        cll.alignment = Alignment(horizontal='center')
                    elif i == 1:
                        cll = sheet.cell(column = i + 1, row = row, value = char.level)
                        cll.alignment = Alignment(horizontal='center')
                    elif i == 2:
                        cll = sheet.cell(column = i + 1, row = row, value = char.gold_donation)
                        cll.alignment = Alignment(horizontal='center')
                        clan_donation = clan.gold_donation
                        mem_donation_total = clan_donation
                        if char.advanced_gold != 0:
                            char.gold_donation += char.advanced_gold 
                        
                        if char.gold_debt > 0:
                            mem_donation_total += char.gold_debt

                        if char.gold_donation < mem_donation_total and not char.gold_donation == 0:
                            cll.fill = PatternFill(start_color="b0e600", fill_type = "solid")
                            continue
                        
                        elif char.gold_donation == 0:
                            cll.fill = PatternFill(start_color="9c0505", fill_type = "solid")
                            continue

                        elif char.gold_donation == mem_donation_total:
                            cll.fill = PatternFill(start_color="1adb14", fill_type = "solid")
                            continue

                        elif char.gold_donation > mem_donation_total:
                            cll.fill = PatternFill(start_color="5b08c7", fill_type = "solid")
                            continue

                    elif i == 3:
                        cll = sheet.cell(column = i + 1, row = row, value = char.gold_debt)
                        cll.alignment = Alignment(horizontal='center')
                    elif i == 4:
                        cll = sheet.cell(column = i + 1, row = row, value = char.advanced_gold)
                        cll.alignment = Alignment(horizontal='center')
                row += 1

            for column_cells in sheet.columns: 
                unmerged_cells = list(filter(lambda cell_to_check: cell_to_check.coordinate not in sheet.merged_cells, column_cells)) 
                length = max(len(str(cell.value)) for cell in unmerged_cells) 
                sheet.column_dimensions[unmerged_cells[0].column_letter].width = length * 1.2
            
            workbook.save(filename=filepath)

            with open(filepath, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = 'attachment; filename=' + filepath
                response['Content-Type'] = 'application/vnd.ms-excel; charset=utf-16'
                return response
            
        else:
            form = ClanForm()
            return render(request, 'admin/excel_report.html', {'form': form, 'title': u'Create excel report for which clan?', 'site_header': "Generate Clan Excel Report"})
    
    create_report.label = "Create Excel Report"
    create_report.short_description = "Generates excel file for saving."